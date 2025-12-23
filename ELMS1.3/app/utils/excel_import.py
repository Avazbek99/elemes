from flask import flash
from app.models import Subject, Faculty
from app import db
from datetime import datetime
import io
import re


def generate_sample_file():
    """Talabalarni import qilish uchun namuna Excel fayl (yangi tartib)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar import"

    # Sarlavha
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = "Talabalar import uchun namuna fayl"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Import talablari
    from datetime import datetime
    ws['A2'] = "IMPORT TALABLARI:"
    ws.merge_cells('A2:O2')
    ws['A2'].font = Font(size=11, bold=True, color="000000")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    
    # Talablar ro'yxati
    requirements = [
        "1. Talaba ID - ixtiyoriy maydon, unikal bo'lishi kerak",
        "2. To'liq ism - majburiy maydon",
        "3. Pasport seriya raqami - majburiy maydon (masalan: AB1234567)",
        "4. JSHSHIR - ixtiyoriy maydon (14 raqam)",
        "5. Tug'ilgan sana - ixtiyoriy maydon (DD.MM.YYYY yoki YYYY-MM-DD formatida)",
        "6. Telefon - ixtiyoriy maydon",
        "7. Email - ixtiyoriy maydon, unikal bo'lishi kerak",
        "8. Tavsif - ixtiyoriy maydon",
        "9. Fakultet - guruh biriktirish uchun kerak",
        "10. Kurs - guruh biriktirish uchun kerak (1-kurs, 2-kurs formatida)",
        "11. Semestr - ixtiyoriy maydon (1-semestr, 2-semestr formatida)",
        "12. Ta'lim shakli - guruh biriktirish uchun kerak (Kunduzgi, Sirtqi, Kechki - bosh harf katta)",
        "13. Mutaxassislik kodi - ixtiyoriy maydon (yo'nalish kodi)",
        "14. Mutaxassislik nomi - ixtiyoriy maydon (yo'nalish nomi)",
        "15. Guruh - guruh biriktirish uchun kerak (agar mavjud bo'lsa, qo'shiladi, aks holda yangi yaratiladi)"
    ]
    
    for idx, req in enumerate(requirements, start=3):
        ws.merge_cells(f'A{idx}:O{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = req
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = [
        "Talaba ID",              # A
        "To'liq ism",             # B
        "Pasport seriya raqami",  # C
        "JSHSHIR",                # D
        "Tug'ilgan sana",         # E
        "Telefon",                # F
        "Email",                  # G
        "Tavsif",                 # H
        "Fakultet",               # I
        "Kurs",                   # J
        "Semestr",                # K
        "Ta'lim shakli",          # L
        "Mutaxassislik kodi",     # M
        "Mutaxassislik nomi",     # N
        "Guruh"                   # O
    ]

    header_row = len(requirements) + 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Namuna ma'lumotlar
    sample_data = [
        ["ST2024001", "Aliyev Vali", "AB1234567", "30202020200021", "15.01.2000", "+998901234567", "vali@example.com", "Talaba haqida ma'lumot", "IT", "1-kurs", "1-semestr", "Kunduzgi", "DI", "Dasturiy injiniring", "DI-21"],
        ["ST2024002", "Karimova Zuhra", "AC2345678", "30202020200022", "20.03.2001", "+998901234568", "zuhra@example.com", "Talaba haqida ma'lumot", "IT", "1-kurs", "1-semestr", "Kunduzgi", "DI", "Dasturiy injiniring", "DI-21"]
    ]

    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Ustun kengliklarini sozlash
    column_widths = [15, 30, 20, 18, 18, 16, 25, 40, 20, 12, 12, 15, 20, 30, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_staff_sample_file():
    """Xodimlarni import qilish uchun namuna Excel fayl (bitta sheet'da)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Xodimlar"
    
    # Sarlavha
    title = "Xodimlar import uchun namuna fayl"
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Import talablari
    from datetime import datetime
    ws['A2'] = "IMPORT TALABLARI:"
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(size=11, bold=True, color="000000")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    
    # Talablar ro'yxati
    requirements = [
        "1. To'liq ism - majburiy maydon",
        "2. Login - majburiy maydon, unikal bo'lishi kerak",
        "3. Pasport seriya raqami - majburiy maydon (masalan: AB1234567)",
        "4. JSHSHIR - ixtiyoriy maydon (14 raqam)",
        "5. Tug'ilgan sana - ixtiyoriy maydon (DD.MM.YYYY yoki YYYY-MM-DD formatida)",
        "6. Telefon - ixtiyoriy maydon",
        "7. Email - ixtiyoriy maydon, unikal bo'lishi kerak",
        "8. Tavsif - ixtiyoriy maydon"
    ]
    
    for idx, req in enumerate(requirements, start=3):
        ws.merge_cells(f'A{idx}:H{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = req
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    
    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = ["To'liq ism", 'Login', 'Pasport seriya raqami', 'JSHSHIR', "Tug'ilgan sana", 'Telefon', 'Email', 'Tavsif']
    header_row = len(requirements) + 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Namuna ma'lumotlar
    sample_data = [
        ["Tursunqulov Avazbek", "admin", "AB1234567", "30202020200021", "15.01.1980", "+998901234567", "admin@university.uz", "Tizim administratori"],
        ["Karimov Sherzod", "sherzod", "AC2345678", "30202020200022", "20.03.1975", "+998901234568", "dean.it@university.uz", "IT fakulteti dekani"],
        ["Mamatov Valijon", "valijon", "AD3456789", "30202020200023", "10.05.1985", "+998901234569", "valijon@university.uz", "Dasturiy injiniring kafedrasi o'qituvchisi"],
        ["Rahimova Aziza", "aziza", "AE4567890", "30202020200024", "25.07.1990", "+998901234570", "accounting@university.uz", "Buxgalteriya bo'limi xodimi"],
        ["Aliyev Vali", "vali", "AF5678901", "30202020200025", "30.09.1982", "+998901234571", "vali@university.uz", "IT fakulteti dekani va o'qituvchi"]
    ]
    
    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Ustun kengliklarini sozlash
    column_widths = [30, 20, 20, 18, 18, 16, 25, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_students_from_excel(file, faculty_id=None):
    """Excel fayldan talabalarni import qilish (yangi tartib)
    
    Args:
        file: Excel fayl
        faculty_id: Fakultet ID (ixtiyoriy, agar berilsa, guruhlar shu fakultet doirasida qidiriladi)
    """
    try:
        from openpyxl import load_workbook
        from app.models import User, Group, Faculty, Direction
        from app import db
        from datetime import datetime, date
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        updated = 0
        errors = []
        
        # Sarlavha qatorini topish (dinamik ravishda)
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            first_cell = ws.cell(row=row_num, column=1).value
            if first_cell and ("Talaba ID" in str(first_cell) or "To'liq ism" in str(first_cell)):
                header_row = row_num
                break
        
        if not header_row:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': ["Sarlavha qatori topilmadi. Iltimos, fayl formati to'g'ri ekanligini tekshiring."]
            }
        
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get("To'liq ism") and not row_data.get('Email'):
                    continue
                
                full_name = row_data.get("To'liq ism", '').strip()
                student_id = row_data.get('Talaba ID', '').strip()
                email = row_data.get('Email', '').strip()
                passport_number = row_data.get('Pasport seriya raqami', '').strip()
                
                if not full_name or not passport_number:
                    errors.append(f"Qator {row_num}: To'liq ism yoki pasport seriya raqami kiritilmagan")
                    continue
                
                # Foydalanuvchini topish (student_id, email yoki passport_number orqali)
                user = None
                if student_id:
                    user = User.query.filter_by(student_id=student_id).first()
                if not user and email:
                    user = User.query.filter_by(email=email).first()
                if not user and passport_number:
                    user = User.query.filter_by(passport_number=passport_number).first()
                
                # JSHSHIR
                pinfl = row_data.get('JSHSHIR', '').strip() or None
                
                # Tug'ilgan sana
                birth_date_str = row_data.get("Tug'ilgan sana", '').strip()
                birth_date = None
                if birth_date_str:
                    try:
                        # DD.MM.YYYY yoki YYYY-MM-DD formatini qo'llab-quvvatlash
                        if '.' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y').date()
                        elif '-' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format (DD.MM.YYYY yoki YYYY-MM-DD)")
                
                # Fakultet, Kurs, Semestr, Ta'lim shakli, Mutaxassislik, Guruh
                faculty_name = row_data.get('Fakultet', '').strip()
                course_str = row_data.get('Kurs', '').strip()  # "1-kurs" formatida
                semester_str = row_data.get('Semestr', '').strip()
                education_type = row_data.get("Ta'lim shakli", '').strip()
                specialty_code = row_data.get('Mutaxassislik kodi', '').strip()
                specialty_name = row_data.get('Mutaxassislik nomi', '').strip()
                group_name = row_data.get('Guruh', '').strip()
                
                # Kurs raqamini ajratish ("1-kurs" -> 1)
                course_year = None
                if course_str:
                    try:
                        course_year = int(course_str.replace('-kurs', '').strip())
                    except:
                        pass
                
                # Semestr raqamini ajratish ("1-semestr" -> 1)
                semester = None
                if semester_str:
                    try:
                        # "1-semestr" formatidan raqamni ajratish
                        semester_str_clean = semester_str.replace('-semestr', '').strip()
                        semester = int(semester_str_clean)
                    except:
                        pass
                
                # Ta'lim shaklini kichik harfga o'tkazish (database'da kichik harfda saqlanadi)
                if education_type:
                    education_type = education_type.lower()
                
                # Yo'nalishni topish yoki yaratish
                direction = None
                if specialty_code and faculty_name:
                    # Fakultetni topish
                    faculty = Faculty.query.filter_by(name=faculty_name).first()
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                        continue
                    
                    # Yo'nalishni topish (kodi bo'yicha)
                    direction = Direction.query.filter_by(code=specialty_code, faculty_id=faculty.id).first()
                    
                    # Agar yo'nalish topilmasa, yangi yaratish
                    if not direction:
                        if specialty_name and course_year and semester:
                            direction = Direction(
                                name=specialty_name,
                                code=specialty_code,
                                faculty_id=faculty.id,
                                course_year=course_year,
                                semester=semester,
                                education_type=education_type or 'kunduzgi'
                            )
                            db.session.add(direction)
                            db.session.flush()
                
                # Guruhni topish yoki yaratish
                group = None
                if group_name and faculty_name:
                    # Fakultetni topish (agar yo'nalishda topilmagan bo'lsa)
                    if not direction:
                        faculty = Faculty.query.filter_by(name=faculty_name).first()
                        if not faculty:
                            errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                            continue
                    else:
                        faculty = direction.faculty
                    
                    # Guruhni topish
                    group = Group.query.filter_by(name=group_name, faculty_id=faculty.id).first()
                    
                    # Agar guruh topilmasa, yangi yaratish
                    if not group:
                        if course_year and education_type:
                            group = Group(
                                name=group_name,
                                faculty_id=faculty.id,
                                direction_id=direction.id if direction else None,
                                course_year=course_year,
                                education_type=education_type
                            )
                            db.session.add(group)
                            db.session.flush()
                        else:
                            errors.append(f"Qator {row_num}: Guruh yaratish uchun kurs va ta'lim shakli kerak")
                    elif direction and not group.direction_id:
                        # Agar guruh topilgan bo'lsa va yo'nalish biriktirilmagan bo'lsa, biriktirish
                        group.direction_id = direction.id
                
                if user:
                    # Yangilash
                    user.full_name = full_name
                    if student_id:
                        user.student_id = student_id
                    user.phone = row_data.get('Telefon', '').strip() or None
                    user.passport_number = passport_number
                    user.pinfl = pinfl
                    user.birth_date = birth_date
                    user.email = email if email else None
                    user.description = row_data.get('Tavsif', '').strip() or None
                    
                    # Guruhni biriktirish
                    if group:
                        user.group_id = group.id
                        if semester:
                            user.semester = semester
                        if education_type:
                            user.education_type = education_type
                    
                    user.set_password(passport_number)
                    updated += 1
                else:
                    # Yaratish
                    user = User(
                        full_name=full_name,
                        role='student',
                        student_id=student_id or None,
                        phone=row_data.get('Telefon', '').strip() or None,
                        passport_number=passport_number,
                        pinfl=pinfl,
                        birth_date=birth_date,
                        email=email if email else None,
                        description=row_data.get('Tavsif', '').strip() or None,
                        semester=semester,
                        education_type=education_type if education_type else None
                    )
                    
                    # Guruhni biriktirish
                    if group:
                        user.group_id = group.id
                    
                    user.set_password(passport_number)
                    db.session.add(user)
                    
                    # Commit qilish va agar email NOT NULL xatolik bo'lsa, email maydonini bo'sh qatorga o'zgartirish
                    try:
                        db.session.flush()  # ID olish uchun
                    except Exception as e:
                        error_str = str(e).lower()
                        if 'email' in error_str and ('not null' in error_str or 'constraint' in error_str):
                            # Database'da email NOT NULL bo'lsa, bo'sh qator qo'yamiz
                            db.session.rollback()
                            user.email = ''  # Bo'sh qator (database NOT NULL constraint uchun)
                            db.session.add(user)
                            db.session.flush()
                        else:
                            raise
                    
                    imported += 1
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_directions_from_excel(file):
    """Excel fayldan yo'nalishlar va guruhlarni import qilish"""
    try:
        from openpyxl import load_workbook
        from app.models import Direction, Group, Faculty
        from app import db
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        errors = []
        
        # Sarlavha qatorini topish (1-qator)
        header_row = 1
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get('Yo\'nalish nomi') and not row_data.get('Yo\'nalish kodi'):
                    continue
                
                direction_name = row_data.get('Yo\'nalish nomi', '').strip()
                direction_code = row_data.get('Yo\'nalish kodi', '').strip()
                faculty_name = row_data.get('Fakultet', '').strip()
                group_name = row_data.get('Guruh', '').strip()
                course_year = row_data.get('Kurs', '').strip()
                
                if not direction_name or not direction_code:
                    errors.append(f"Qator {row_num}: Yo'nalish nomi yoki kodi kiritilmagan")
                    continue
                
                # Fakultetni topish
                faculty = None
                if faculty_name:
                    faculty = Faculty.query.filter_by(name=faculty_name).first()
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                        continue
                
                # Yo'nalishni topish yoki yaratish
                direction = Direction.query.filter_by(code=direction_code).first()
                if not direction:
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet kiritilmagan")
                        continue
                    direction = Direction(
                        name=direction_name,
                        code=direction_code,
                        description=row_data.get('Tavsif', '').strip() or None,
                        faculty_id=faculty.id
                    )
                    db.session.add(direction)
                    db.session.flush()
                    imported += 1
                
                # Guruhni topish yoki yaratish
                if group_name and faculty:
                    group = Group.query.filter_by(name=group_name).first()
                    if not group:
                        try:
                            course_year_int = int(course_year) if course_year else 1
                        except:
                            course_year_int = 1
                        
                        group = Group(
                            name=group_name,
                            faculty_id=faculty.id,
                            direction_id=direction.id,
                            course_year=course_year_int
                        )
                        db.session.add(group)
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_staff_from_excel(file):
    """Excel fayldan xodimlarni import qilish (bitta sheet'dan) - bir nechta rollarni qo'llab-quvvatlash"""
    try:
        from openpyxl import load_workbook
        from app.models import User, Faculty, UserRole
        from app import db
        from datetime import datetime, date
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active  # Bitta sheet'dan o'qish
        
        imported = 0
        updated = 0
        errors = []
        
        # Sarlavha qatorini topish (dinamik ravishda)
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):  # Birinchi 20 qatorni tekshirish
            first_cell = ws.cell(row=row_num, column=1).value
            if first_cell and ("To'liq ism" in str(first_cell) or "To'liq ismi" in str(first_cell)):
                header_row = row_num
                break
        
        if not header_row:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': ["Sarlavha qatori topilmadi. Iltimos, fayl formati to'g'ri ekanligini tekshiring."]
            }
        
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get("To'liq ism") and not row_data.get('Email'):
                    continue
                
                full_name = row_data.get("To'liq ism", '').strip()
                login = row_data.get('Login', '').strip()
                email = row_data.get('Email', '').strip()
                
                if not full_name:
                    errors.append(f"Qator {row_num}: To'liq ism kiritilmagan")
                    continue
                
                # Login yoki email orqali foydalanuvchini topish
                user = None
                if login:
                    user = User.query.filter_by(login=login).first()
                if not user and email:
                    user = User.query.filter_by(email=email).first()
                
                # Xodim uchun
                passport_number = row_data.get('Pasport seriya raqami', '').strip()
                if not passport_number:
                    errors.append(f"Qator {row_num}: Pasport seriya raqami kiritilmagan")
                    continue
                
                # Pasport raqamini katta harfga o'zgartirish
                passport_number = passport_number.upper()
                
                # JSHSHIR
                pinfl = row_data.get('JSHSHIR', '').strip() or None
                
                # Tug'ilgan sana
                birth_date_str = row_data.get("Tug'ilgan sana", '').strip()
                birth_date = None
                if birth_date_str:
                    try:
                        # DD.MM.YYYY yoki YYYY-MM-DD formatini qo'llab-quvvatlash
                        if '.' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y').date()
                        elif '-' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                        else:
                            errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format (DD.MM.YYYY yoki YYYY-MM-DD)")
                    except ValueError:
                        errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format")
                
                # Asosiy rol - import qilishda default 'teacher' rolini beramiz
                # (Rollar keyinchalik admin panel orqali belgilanadi)
                primary_role = 'teacher'
                
                if user:
                    # Yangilash
                    user.full_name = full_name
                    if login:
                        user.login = login
                    user.phone = row_data.get('Telefon', '').strip() or None
                    user.passport_number = passport_number
                    user.pinfl = pinfl
                    user.birth_date = birth_date
                    user.email = email if email else None
                    user.description = row_data.get('Tavsif', '').strip() or None
                    
                    user.set_password(passport_number)
                    
                    # Agar foydalanuvchi talaba bo'lsa, rolini o'zgartirish
                    if user.role == 'student':
                        user.role = primary_role
                    
                    updated += 1
                else:
                    # Yaratish
                    # Login majburiy
                    if not login:
                        errors.append(f"Qator {row_num}: Login kiritilmagan")
                        continue
                    
                    # Login unikalligi
                    if User.query.filter_by(login=login).first():
                        errors.append(f"Qator {row_num}: Bu login allaqachon mavjud")
                        continue
                    
                    # Email unikalligi (agar berilsa)
                    if email and User.query.filter_by(email=email).first():
                        errors.append(f"Qator {row_num}: Bu email allaqachon mavjud")
                        continue
                    
                    user = User(
                        login=login,
                        full_name=full_name,
                        role=primary_role,
                        phone=row_data.get('Telefon', '').strip() or None,
                        passport_number=passport_number,
                        pinfl=pinfl,
                        birth_date=birth_date,
                        description=row_data.get('Tavsif', '').strip() or None
                    )
                    
                    # Email maydonini alohida o'rnatish (agar bo'sh bo'lsa, o'rnatmaymiz)
                    if email:
                        user.email = email
                    
                    # Parolni pasport raqamiga o'rnatish
                    user.set_password(passport_number)
                    
                    db.session.add(user)
                    
                    # Commit qilish va agar email NOT NULL xatolik bo'lsa, email maydonini bo'sh qatorga o'zgartirish
                    try:
                        db.session.flush()  # ID olish uchun
                    except Exception as e:
                        error_str = str(e).lower()
                        if 'email' in error_str and ('not null' in error_str or 'constraint' in error_str):
                            # Database'da email NOT NULL bo'lsa, bo'sh qator qo'yamiz
                            db.session.rollback()
                            user.email = ''  # Bo'sh qator (database NOT NULL constraint uchun)
                            db.session.add(user)
                            db.session.flush()
                        else:
                            raise
                    
                    imported += 1
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_all_users_from_excel(file):
    """Excel fayldan barcha foydalanuvchilarni import qilish (rol bo'yicha ajratish) - eski funksiya, import_staff_from_excel ishlatiladi"""
    return import_staff_from_excel(file)


def generate_subjects_sample_file():
    """Fanlarni import qilish uchun namuna Excel fayl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Fanlar import"

    # Sarlavha
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "Fanlar import uchun namuna fayl"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Jadval sarlavhalari
    headers = [
        "Fan nomi",      # A
        "Fan kodi",      # B
        "Tavsif"         # C
    ]

    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Namuna ma'lumotlar
    sample_data = [
        ["Dasturlash asoslari", "DA101", "Dasturlashning asosiy tushunchalari va algoritmlar"],
        ["Ma'lumotlar bazasi", "MB201", "Ma'lumotlar bazasi dizayni va SQL so'rovlari"],
        ["Web dasturlash", "WD301", "Web texnologiyalari va frameworklar"]
    ]

    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Ustun kengliklarini sozlash
    column_widths = [30, 15, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_subjects_from_excel(file):
    """Excel fayldan fanlarni import qilish"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    imported = 0
    updated = 0
    errors = []

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Sarlavha qatorini topish (3-qator)
        header_row = 3
        headers = {}
        for col_num in range(1, 4):  # A, B, C ustunlari
            cell_value = ws.cell(row=header_row, column=col_num).value
            if cell_value:
                headers[cell_value] = col_num

        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                # Ustunlardan ma'lumotlarni olish
                name = ws.cell(row=row_num, column=headers.get("Fan nomi", 1)).value
                code = ws.cell(row=row_num, column=headers.get("Fan kodi", 2)).value
                description = ws.cell(row=row_num, column=headers.get("Tavsif", 3)).value

                # Bo'sh qatorlarni o'tkazib yuborish
                if not name or not code:
                    continue

                name = str(name).strip()
                code = str(code).strip().upper()
                description = str(description).strip() if description else None

                # Fan kodini tekshirish
                existing_subject = Subject.query.filter_by(code=code).first()

                # Default fakultetni tanlash
                default_faculty = Faculty.query.first()
                if not default_faculty:
                    errors.append(f"Qator {row_num}: Fakultet mavjud emas")
                    continue

                if existing_subject:
                    # Yangilash
                    existing_subject.name = name
                    existing_subject.description = description
                    updated += 1
                else:
                    # Yaratish
                    subject = Subject(
                        name=name,
                        code=code,
                        description=description,
                        credits=3,  # Default
                        faculty_id=default_faculty.id,  # Default fakultet
                        semester=1  # Default
                    )
                    db.session.add(subject)
                    imported += 1

            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")

        db.session.commit()

        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }

    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }
